import csv, io, os, sqlite3
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB_PATH = os.environ.get("DB_PATH", "/tmp/contatti.db")
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "residenze2026")
HEADERS = ["Data","Nome","Cognome","Email","Telefono","Villa","Capitolato","Messaggio"]
COL_WIDTHS = [18, 16, 16, 30, 16, 18, 18, 40]

db = sqlite3.connect(DB_PATH, check_same_thread=False)
db.row_factory = sqlite3.Row
db.execute("CREATE TABLE IF NOT EXISTS contatti (id INTEGER PRIMARY KEY AUTOINCREMENT, data TEXT, nome TEXT, cognome TEXT, email TEXT, telefono TEXT, villa TEXT, capitolato TEXT, messaggio TEXT)")
db.commit()

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

class ContactForm(BaseModel):
    nome: str
    cognome: str
    email: str
    telefono: Optional[str] = ""
    villa: Optional[str] = ""
    capitolato: Optional[str] = ""
    messaggio: Optional[str] = ""

@app.post("/api/contact")
async def contact(data: ContactForm):
    db.execute("INSERT INTO contatti (data,nome,cognome,email,telefono,villa,capitolato,messaggio) VALUES (?,?,?,?,?,?,?,?)",
        (datetime.now().strftime("%Y-%m-%d %H:%M"), data.nome, data.cognome, data.email, data.telefono, data.villa, data.capitolato, data.messaggio))
    db.commit()
    return {"ok": True}

@app.get("/api/health")
def health():
    return {"status": "ok", "contatti": db.execute("SELECT COUNT(*) FROM contatti").fetchone()[0]}

@app.get("/api/contacts/export")
def export_csv(token: str = ""):
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    rows = db.execute("SELECT data,nome,cognome,email,telefono,villa,capitolato,messaggio FROM contatti ORDER BY id").fetchall()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(HEADERS)
    for r in rows:
        w.writerow(list(r))
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=contatti.csv"})

@app.get("/api/contacts/export/excel")
def export_excel(token: str = ""):
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    rows = db.execute("SELECT data,nome,cognome,email,telefono,villa,capitolato,messaggio FROM contatti ORDER BY id").fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contatti"
    header_fill = PatternFill(start_color="1A0F00", end_color="1A0F00", fill_type="solid")
    header_font = Font(color="C49A2A", bold=True, name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i-1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    row_fill_even = PatternFill(start_color="FDF6E3", end_color="FDF6E3", fill_type="solid")
    row_fill_odd  = PatternFill(start_color="F5E9CC", end_color="F5E9CC", fill_type="solid")
    for row_idx, row in enumerate(rows, 2):
        fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        for col_idx, value in enumerate(list(row), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contatti_residenze.xlsx"}
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))
